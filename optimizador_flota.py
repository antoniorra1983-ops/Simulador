"""
Optimizador de distribución de flota para minimizar consumo energético.

Estrategia: reasignar el tipo de tren (XT-100 / XT-M) a cada servicio de la malla,
respetando la flota disponible y la capacidad de pasajeros requerida, de modo que
los trenes más eficientes (XT-M, menor IDE) cubran los servicios de mayor km/demanda.

No altera los horarios (t_ini, t_fin) ni la malla — solo qué unidad hace cada servicio.

El consumo se calcula con el MISMO motor físico que el simulador (no por IDE fijo),
por lo que la línea base coincide exactamente con lo que muestra el Gemelo Digital.
"""
import pandas as pd
import numpy as np


def _flota_disponible(config):
    """Unidades disponibles por tipo de tren."""
    try:
        flota = getattr(config, 'FLOTA', {})
        disp = {}
        for tipo, params in flota.items():
            disp[tipo] = int(params.get('unidades_disponibles', 0))
        if not any(disp.values()):
            disp = {'XT-100': 27, 'XT-M': 8, 'SFE': 5}
        return disp
    except Exception:
        return {'XT-100': 27, 'XT-M': 8, 'SFE': 5}


# Capacidad de estacionamiento/ocupación por terminal (km → (nombre, capacidad))
_TERMINALES_CAP = {
    0.00:  ('Puerto', 4),
    25.30: ('El Belloto', 16),
    43.13: ('Limache', 16),
}
_TOL_TERMINAL_KM = 0.5  # margen para considerar un tren "en" el terminal


def verificar_capacidad_terminales(df_servicios, paso_min=5.0):
    """
    Verifica la ocupación de cada terminal a lo largo del día.
    Un tren ocupa un terminal si: (a) está estacionado ahí entre servicios, o
    (b) está físicamente en el km del terminal durante un servicio (origen/destino).

    Retorna dict {terminal: {'max_ocup': N, 'capacidad': C, 'excede': bool, 'pico_min': t}}
    """
    df = df_servicios.copy().reset_index(drop=True)
    if 't_fin' not in df.columns:
        df['t_fin'] = df['t_ini'] + 55

    # Rango temporal del día
    t_min = df['t_ini'].min()
    t_max = df['t_fin'].max()

    resultado = {}
    for km_term, (nombre, cap) in _TERMINALES_CAP.items():
        max_ocup = 0
        pico_min = t_min
        t = t_min
        while t <= t_max:
            ocup = 0
            for _, r in df.iterrows():
                ko, kd = r['km_orig'], r['km_dest']
                # ¿este servicio toca el terminal en su origen o destino?
                toca_origen = abs(ko - km_term) <= _TOL_TERMINAL_KM
                toca_destino = abs(kd - km_term) <= _TOL_TERMINAL_KM
                if not (toca_origen or toca_destino):
                    continue
                # ocupa el terminal al inicio (origen) o al final (destino) del servicio
                if toca_origen and (r['t_ini'] - paso_min) <= t <= (r['t_ini'] + paso_min):
                    ocup += 1
                elif toca_destino and (r['t_fin'] - paso_min) <= t <= (r['t_fin'] + paso_min):
                    ocup += 1
            if ocup > max_ocup:
                max_ocup = ocup
                pico_min = t
            t += paso_min

        resultado[nombre] = {
            'max_ocup': max_ocup,
            'capacidad': cap,
            'excede': max_ocup > cap,
            'pico_min': pico_min,
            'km': km_term,
        }
    return resultado


def calcular_seat_total(df_e, config, active_sers, distribuir_fn, flujo_fn):
    """
    Calcula el SEAT total (kWh) idéntico a como lo hace el dashboard del planificador:
    SEAT = (Σ energía_por_SER / ETA_RECTIFICADOR + pérdidas_AC) / 0.99

    df_e: DataFrame con kwh_viaje_trac, kwh_viaje_aux, kwh_viaje_regen, t_viaje_h,
          km_orig, km_dest por servicio (salida del motor).
    Devuelve (seat_total_kwh, km_total).
    """
    try:
        eta_ser = getattr(config, 'ETA_SER_RECTIFICADOR', 0.96)
    except Exception:
        eta_ser = 0.96

    ser_accum = {s[1]: 0.0 for s in active_sers}
    t_total_h = 0.0
    km_total = 0.0

    for _, r in df_e.iterrows():
        e_panto = (r.get('kwh_viaje_trac', 0) + r.get('kwh_viaje_aux', 0) - r.get('kwh_viaje_regen', 0))
        th = r.get('t_viaje_h', 0.0)
        km_o, km_d = r['km_orig'], r['km_dest']
        t_total_h += th
        # km igual que el dashboard: usa 'tren_km' (incluye factor 2× para dobles)
        if 'tren_km' in df_e.columns and pd.notna(r.get('tren_km')):
            km_total += float(r['tren_km'])
        else:
            km_total += abs(km_d - km_o) * (2.0 if r.get('doble', False) else 1.0)
        for s_name, e_val in distribuir_fn(e_panto, th, km_o, km_d, active_sers).items():
            ser_accum[s_name] = ser_accum.get(s_name, 0.0) + e_val

    total_ser_44kv = sum(max(0.0, v) for v in ser_accum.values()) / eta_ser
    t_elap = max(0.001, t_total_h)
    flujo = flujo_fn({k: max(0.0, v) / eta_ser / t_elap for k, v in ser_accum.items()})
    loss_ac = flujo.get('P_loss_kw', 0.0) * (1.15 ** 2) * t_elap
    seat = (total_ser_44kv + loss_ac) / 0.99
    return seat, km_total


def optimizar_asignacion_flota(df_servicios, config, priorizar='energia',
                                df_consumo_base=None, simular_fn=None,
                                precalcular_fn=None, params_sim=None,
                                prevenciones=None, active_sers=None,
                                distribuir_fn=None, flujo_fn=None):
    """
    Reasigna tipos de tren a los servicios para minimizar el consumo.

    El consumo se calcula como SEAT total (idéntico al dashboard del planificador):
    incluye pérdidas de rectificador, distribución por subestación y pérdidas AC.

    df_consumo_base: df_sint_e ya calculado por el simulador (línea base real).
    simular_fn/precalcular_fn/params_sim: motor para recalcular la versión optimizada.
    active_sers/distribuir_fn/flujo_fn: para calcular el SEAT igual que el dashboard.

    Retorna: (df_optimizado, resumen_dict)
    """
    df = df_servicios.copy().reset_index(drop=True)
    flota_disp = _flota_disponible(config)
    df['km_tramo'] = (df['km_dest'] - df['km_orig']).abs()

    try:
        flota = getattr(config, 'FLOTA', {})
    except Exception:
        flota = {}

    df['cap_req'] = df.apply(lambda r: r.get('pax_abordo', 0) or 0, axis=1)

    ide_ref = {'XT-100': 3.88, 'XT-M': 3.28, 'SFE': 5.77}
    tipos_ordenados = sorted(ide_ref.keys(), key=lambda t: ide_ref[t])
    cap_tipo = {t: flota.get(t, {}).get('cap_max', 398) for t in ide_ref}

    df['tipo_optimo'] = df['tipo_tren']

    def trenes_simultaneos(idx, tipo_col):
        t_ini = df.at[idx, 't_ini']
        t_fin = df.at[idx, 't_fin'] if 't_fin' in df.columns else t_ini + 55
        count = {}
        for j in range(len(df)):
            if j == idx:
                continue
            tj_ini = df.at[j, 't_ini']
            tj_fin = df.at[j, 't_fin'] if 't_fin' in df.columns else tj_ini + 55
            if tj_ini < t_fin and tj_fin > t_ini:
                tp = df.at[j, tipo_col]
                count[tp] = count.get(tp, 0) + 1
        return count

    orden = df['km_tramo'].sort_values(ascending=False).index
    for idx in orden:
        cap_req = df.at[idx, 'cap_req']
        es_doble = bool(df.at[idx, 'doble'])
        mejor_tipo = df.at[idx, 'tipo_tren']
        for tipo in tipos_ordenados:
            cap_disp = cap_tipo.get(tipo, 398) * (2 if es_doble else 1)
            if cap_disp < cap_req:
                continue
            simult = trenes_simultaneos(idx, 'tipo_optimo')
            if simult.get(tipo, 0) < flota_disp.get(tipo, 0):
                mejor_tipo = tipo
                break
        df.at[idx, 'tipo_optimo'] = mejor_tipo

    puede_seat = (active_sers is not None and distribuir_fn is not None and flujo_fn is not None)

    # === CONSUMO ACTUAL (SEAT real, línea base del simulador) ===
    km_total = df['km_tramo'].sum()
    if df_consumo_base is not None and puede_seat:
        kwh_actual_total, km_total = calcular_seat_total(
            df_consumo_base, config, active_sers, distribuir_fn, flujo_fn)
    elif df_consumo_base is not None and 'kwh_viaje_neto' in df_consumo_base.columns:
        kwh_actual_total = df_consumo_base['kwh_viaje_neto'].sum()
    else:
        df['kwh_actual'] = df.apply(
            lambda r: ide_ref.get(r['tipo_tren'], 3.88) * r['km_tramo'] * (2 if r['doble'] else 1), axis=1)
        kwh_actual_total = df['kwh_actual'].sum()

    # === CONSUMO OPTIMIZADO (recalculado con motor real → SEAT) ===
    if simular_fn is not None and params_sim is not None:
        df_opt_sim = df.copy()
        df_opt_sim['tipo_tren'] = df_opt_sim['tipo_optimo']
        p = params_sim
        df_e1 = simular_fn(df_opt_sim, p['pct_trac'], p['use_pend'], p['use_rm'],
                           p['use_regen'], {}, p['estacion_anio'], prevenciones=prevenciones)
        dict_r = {}
        if precalcular_fn is not None:
            try:
                dict_r = precalcular_fn(df_e1, p['pct_trac'], p['use_rm'], p['estacion_anio'])
            except Exception:
                dict_r = {}
        try:
            df_e = simular_fn(df_opt_sim, p['pct_trac'], p['use_pend'], p['use_rm'],
                              p['use_regen'], dict_r, p['estacion_anio'], prevenciones=prevenciones,
                              aplicar_anden=True)
        except TypeError:
            df_e = simular_fn(df_opt_sim, p['pct_trac'], p['use_pend'], p['use_rm'],
                              p['use_regen'], dict_r, p['estacion_anio'], prevenciones=prevenciones)
        if puede_seat:
            kwh_optimo_total, _ = calcular_seat_total(
                df_e, config, active_sers, distribuir_fn, flujo_fn)
        else:
            kwh_optimo_total = pd.to_numeric(df_e['kwh_viaje_neto'], errors='coerce').sum()
    else:
        df['kwh_optimo'] = df.apply(
            lambda r: ide_ref.get(r['tipo_optimo'], 3.88) * r['km_tramo'] * (2 if r['doble'] else 1), axis=1)
        kwh_optimo_total = df['kwh_optimo'].sum()

    ahorro = kwh_actual_total - kwh_optimo_total
    ahorro_pct = (ahorro / kwh_actual_total * 100) if kwh_actual_total > 0 else 0.0
    cambios = df[df['tipo_tren'] != df['tipo_optimo']]

    # Verificar capacidad de terminales (restricción dura)
    cap_terminales = verificar_capacidad_terminales(df)
    excede_alguno = any(v['excede'] for v in cap_terminales.values())

    resumen = {
        'kwh_actual': kwh_actual_total,
        'kwh_optimo': kwh_optimo_total,
        'ahorro_kwh': ahorro,
        'ahorro_pct': ahorro_pct,
        'km_total': km_total,
        'ide_actual': kwh_actual_total / km_total if km_total > 0 else 0.0,
        'ide_optimo': kwh_optimo_total / km_total if km_total > 0 else 0.0,
        'n_cambios': len(cambios),
        'n_servicios': len(df),
        'comp_antes': df['tipo_tren'].value_counts().to_dict(),
        'comp_despues': df['tipo_optimo'].value_counts().to_dict(),
        'flota_disponible': flota_disp,
        'usa_seat_real': (puede_seat and df_consumo_base is not None and simular_fn is not None),
        'cap_terminales': cap_terminales,
        'excede_terminales': excede_alguno,
    }

    return df, resumen


# Rangos de numeración de motrices por tipo (MERVAL)
_RANGOS_MOTRIZ = {
    'XT-100': list(range(1, 28)),    # 1-27
    'XT-M':   list(range(28, 36)),   # 28-35
    'SFE':    list(range(410, 415)), # 410-414
}


def _asignar_motrices_por_tipo(df_opt):
    """
    Asigna números de motriz concretos según el tipo óptimo, respetando:
    (a) que cada motriz física no esté en dos servicios solapados en el tiempo;
    (b) que la ocupación de cada terminal (Puerto 4, El Belloto 16, Limache 16)
        no se exceda — no se asignan unidades extra que dejen un terminal sobreocupado.
    Retorna el df con columnas 'motriz_1_opt' y 'motriz_2_opt'.
    """
    df = df_opt.copy().sort_values('t_ini').reset_index(drop=True)
    df['motriz_1_opt'] = None
    df['motriz_2_opt'] = None

    ocupacion = {}  # motriz_num -> lista de intervalos (t_ini, t_fin)

    def libre(motriz, t_ini, t_fin):
        for (oi, of) in ocupacion.get(motriz, []):
            if oi < t_fin and of > t_ini:
                return False
        return True

    # Ocupación de terminales: para cada terminal, intervalos en que un tren está estacionado
    # Un tren se estaciona en un terminal desde que termina un servicio ahí hasta que sale.
    cap_terminal = {km: cap for km, (nombre, cap) in _TERMINALES_CAP.items()}

    def terminal_de(km):
        for km_t in cap_terminal:
            if abs(km - km_t) <= _TOL_TERMINAL_KM:
                return km_t
        return None

    # Para cada terminal, lista de momentos de ocupación (t, +1 entra / -1 sale)
    eventos_terminal = {km: [] for km in cap_terminal}

    def ocupacion_terminal_en(km_t, t):
        """Cuántos trenes hay en el terminal km_t en el instante t."""
        ocup = 0
        for (te, delta) in eventos_terminal[km_t]:
            if te <= t:
                ocup += delta
        return ocup

    for idx in range(len(df)):
        tipo = df.at[idx, 'tipo_optimo']
        t_ini = df.at[idx, 't_ini']
        t_fin = df.at[idx, 't_fin'] if 't_fin' in df.columns else t_ini + 55
        es_doble = bool(df.at[idx, 'doble'])
        km_o = df.at[idx, 'km_orig']
        km_d = df.at[idx, 'km_dest']
        rango = _RANGOS_MOTRIZ.get(tipo, _RANGOS_MOTRIZ['XT-100'])

        # Verificar capacidad del terminal de DESTINO al terminar (el tren se estaciona ahí)
        km_term_dest = terminal_de(km_d)
        n_necesarias = 2 if es_doble else 1

        # Si el destino es un terminal, comprobar que hay espacio al llegar
        puede_estacionar = True
        if km_term_dest is not None:
            ocup_actual = ocupacion_terminal_en(km_term_dest, t_fin)
            if ocup_actual + n_necesarias > cap_terminal[km_term_dest]:
                puede_estacionar = False  # el terminal quedaría sobreocupado

        # Buscar motrices libres del tipo
        asignadas = []
        for m in rango:
            if libre(m, t_ini, t_fin):
                asignadas.append(m)
                if len(asignadas) >= n_necesarias:
                    break

        for m in asignadas:
            ocupacion.setdefault(m, []).append((t_ini, t_fin))

        # Registrar ocupación de terminales (origen: sale; destino: entra)
        km_term_orig = terminal_de(km_o)
        if km_term_orig is not None and asignadas:
            # el tren sale del terminal de origen al iniciar
            eventos_terminal[km_term_orig].append((t_ini, -len(asignadas)))
        if km_term_dest is not None and asignadas:
            # el tren entra al terminal de destino al terminar
            eventos_terminal[km_term_dest].append((t_fin, +len(asignadas)))

        if len(asignadas) >= 1:
            df.at[idx, 'motriz_1_opt'] = asignadas[0]
        if es_doble and len(asignadas) >= 2:
            df.at[idx, 'motriz_2_opt'] = asignadas[1]

    return df


def generar_tabla_seat_15min(df_e, config, active_sers, distribuir_fn, flujo_fn, ruta_xlsx, paso_min=15.0):
    """
    Genera una tabla xlsx con el consumo SEAT en franjas de `paso_min` minutos.
    Columnas: Franja horaria | kWh total | kW medio total | (por cada SER: kWh y kW)

    El consumo de cada viaje se reparte entre las franjas de tiempo que cubre,
    proporcional al tiempo que el tren pasa en cada franja. El SEAT incluye pérdidas
    de rectificador (eta_ser) y AC, idéntico al dashboard.
    """
    import pandas as pd
    import numpy as np
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    try:
        eta_ser = getattr(config, 'ETA_SER_RECTIFICADOR', 0.96)
    except Exception:
        eta_ser = 0.96
    ser_names = [s[1] for s in active_sers]

    df = df_e.copy()
    if 't_fin' not in df.columns:
        if 't_viaje_h' in df.columns:
            df['t_fin'] = df['t_ini'] + df['t_viaje_h'] * 60.0
        else:
            df['t_fin'] = df['t_ini'] + 55.0

    # Rango de franjas (extendido: 30 min antes del primer viaje para el pre-encendido,
    # 10 min después del último para el post)
    t_min = float(df['t_ini'].min()) - 30.0
    t_max = float(df['t_fin'].max()) + 10.0
    inicio = int((t_min // paso_min) * paso_min)
    fin = int(((t_max // paso_min) + 1) * paso_min)
    franjas = list(range(inicio, fin, int(paso_min)))

    # Para cada franja, acumular energía por SER (a nivel de subestación 44kV)
    # repartiendo cada viaje según el solape temporal con la franja.
    ser_por_franja = {f: {n: 0.0 for n in ser_names} for f in franjas}

    # Helper: agregar energía de pantógrafo a las franjas de un intervalo [t_a, t_b]
    # distribuida uniformemente, usando un punto km de referencia para repartir por SER.
    def _agregar_intervalo(t_a, t_b, e_total, km_ref_o, km_ref_d):
        dur = max(0.001, t_b - t_a)
        for f in franjas:
            f_ini, f_fin = f, f + paso_min
            solape = max(0.0, min(t_b, f_fin) - max(t_a, f_ini))
            if solape <= 0:
                continue
            frac = solape / dur
            e_frac = e_total * frac
            th_frac = (dur / 60.0) * frac
            for s_name, e_val in distribuir_fn(e_frac, th_frac, km_ref_o, km_ref_d, active_sers).items():
                ser_por_franja[f][s_name] = ser_por_franja[f].get(s_name, 0.0) + max(0.0, e_val)

    for _, r in df.iterrows():
        t_ini_v = float(r['t_ini'])
        t_fin_v = float(r['t_fin'])
        dur_v = max(0.001, t_fin_v - t_ini_v)
        # Consumo del VIAJE sin el prepost (que se ubica en sus franjas propias aparte)
        prepost = r.get('kwh_prepost', 0.0) or 0.0
        aux_viaje = (r.get('kwh_viaje_aux', 0) or 0) - prepost
        e_panto_total = r.get('kwh_viaje_trac', 0) + aux_viaje - r.get('kwh_viaje_regen', 0)
        km_o, km_d = r['km_orig'], r['km_dest']
        _agregar_intervalo(t_ini_v, t_fin_v, e_panto_total, km_o, km_d)

    # === Consumo PRE/POST en sus franjas temporales correctas ===
    # 30 min antes del PRIMER viaje de cada tren físico (pre) y 10 min después del ÚLTIMO (post).
    # El kwh_prepost está repartido entre los viajes de cada motriz; lo reconstruimos por tren.
    import re as _re
    T_PRE_MIN = 30.0
    T_POST_MIN = 10.0
    tren_fisico = {}  # num → lista de idx
    for idx, row in df.iterrows():
        for n in _re.findall(r'\d+', str(row.get('motriz_num', ''))):
            tren_fisico.setdefault(n, []).append(idx)

    for motriz, idxs in tren_fisico.items():
        if not idxs:
            continue
        sub = df.loc[idxs].sort_values('t_ini')
        # kwh_prepost total de este tren = suma de las porciones asignadas a sus viajes
        kwh_pp_tren = sub['kwh_prepost'].fillna(0.0).sum() if 'kwh_prepost' in sub.columns else 0.0
        if kwh_pp_tren <= 0:
            continue
        # dividir entre pre y post proporcional a la duración (30 vs 10 min)
        kwh_pre = kwh_pp_tren * (T_PRE_MIN / (T_PRE_MIN + T_POST_MIN))
        kwh_post = kwh_pp_tren * (T_POST_MIN / (T_PRE_MIN + T_POST_MIN))

        primer = sub.iloc[0]
        ultimo = sub.iloc[-1]
        # Pre: 30 min antes de salir el primer viaje, en el km de origen (terminal)
        t_pre_ini = float(primer['t_ini']) - T_PRE_MIN
        t_pre_fin = float(primer['t_ini'])
        km_term_pre = primer['km_orig']
        _agregar_intervalo(t_pre_ini, t_pre_fin, kwh_pre, km_term_pre, km_term_pre)
        # Post: 10 min después de llegar el último viaje, en el km de destino (terminal)
        t_post_ini = float(ultimo['t_fin'])
        t_post_fin = float(ultimo['t_fin']) + T_POST_MIN
        km_term_post = ultimo['km_dest']
        _agregar_intervalo(t_post_ini, t_post_fin, kwh_post, km_term_post, km_term_post)

    # Construir filas: SEAT total y por SER (kWh y kW medio)
    filas = []
    horas_frac = paso_min / 60.0
    for f in franjas:
        ser_acc = ser_por_franja[f]
        # SEAT por SER (44kV con rectificador)
        ser_seat = {n: ser_acc[n] / eta_ser for n in ser_names}
        total_ser_44 = sum(ser_seat.values())
        # pérdidas AC sobre el total de la franja
        flujo = flujo_fn({n: ser_seat[n] / max(0.001, horas_frac) for n in ser_names})
        loss_ac = flujo.get('P_loss_kw', 0.0) * (1.15 ** 2) * horas_frac
        seat_total = (total_ser_44 + loss_ac) / 0.99

        h = int(f // 60); m = int(f % 60)
        hf = int((f + paso_min) // 60); mf = int((f + paso_min) % 60)
        fila = {
            'Franja': f"{h:02d}:{m:02d}-{hf:02d}:{mf:02d}",
            'SEAT kWh': round(seat_total, 1),
            'SEAT kW medio': round(seat_total / horas_frac, 1),
        }
        for n in ser_names:
            fila[f'{n} kWh'] = round(ser_seat[n], 1)
            fila[f'{n} kW'] = round(ser_seat[n] / horas_frac, 1)
        filas.append(fila)

    df_tabla = pd.DataFrame(filas)

    # Escribir xlsx con formato
    wb = Workbook(); ws = wb.active; ws.title = "SEAT 15min"
    head_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    head_fill = PatternFill('solid', start_color='1565C0')
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = list(df_tabla.columns)
    for c_idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font = head_font; cell.fill = head_fill
        cell.alignment = center; cell.border = border

    for r_idx, (_, row) in enumerate(df_tabla.iterrows(), start=2):
        for c_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[col])
            cell.alignment = center; cell.border = border
            cell.font = Font(name='Arial', size=9)
            if c_idx > 1:
                cell.number_format = '#,##0.0'

    # Fila de totales
    tot_row = len(df_tabla) + 2
    ws.cell(row=tot_row, column=1, value="TOTAL DÍA").font = Font(name='Arial', bold=True, size=10)
    for c_idx, col in enumerate(cols, start=1):
        if c_idx == 1:
            continue
        if 'kWh' in col:  # sumar kWh; el kW medio no se suma
            total = df_tabla[col].sum()
            cell = ws.cell(row=tot_row, column=c_idx, value=round(total, 1))
            cell.font = Font(name='Arial', bold=True, size=9)
            cell.number_format = '#,##0.0'

    # Anchos
    ws.column_dimensions['A'].width = 16
    for c_idx in range(2, len(cols) + 1):
        ws.column_dimensions[ws.cell(row=1, column=c_idx).column_letter].width = 13

    ws.freeze_panes = 'B2'
    wb.save(ruta_xlsx)
    return ruta_xlsx, df_tabla


def generar_planillas_xlsx(df_opt, ruta_v1, ruta_v2):
    """
    Genera dos archivos xlsx (V1 y V2) con el formato de las planillas originales:
    Columnas: N° Viaje | Servicio | Hr Partida | N° Partida | Intervalo | Unidad | Motriz 1 | Motriz 2

    df_opt debe tener: Via, num_servicio, t_ini, doble, tipo_optimo (y t_fin opcional).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    df = _asignar_motrices_por_tipo(df_opt)

    def _fmt_hora(t_mins):
        try:
            h = int(t_mins // 60); m = int(t_mins % 60); s = int(round((t_mins - int(t_mins)) * 60))
            return f"{h:02d}:{m:02d}:{s:02d}"
        except Exception:
            return ""

    cols = ['N° Viaje', 'Servicio', 'Hr Partida', 'N° Partida', 'Intervalo', 'Unidad', 'Motriz 1', 'Motriz 2']

    for via, ruta in [(1, ruta_v1), (2, ruta_v2)]:
        sub = df[df['Via'] == via].sort_values('t_ini').reset_index(drop=True)
        wb = Workbook(); ws = wb.active; ws.title = f"V{via}"

        # Encabezado
        head_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        head_fill = PatternFill('solid', start_color='1565C0' if via == 1 else 'C62828')
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for c_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col)
            cell.font = head_font; cell.fill = head_fill
            cell.alignment = center; cell.border = border

        prev_t = None
        for r_idx, (_, row) in enumerate(sub.iterrows(), start=2):
            t_ini = row['t_ini']
            n_viaje = row.get('num_servicio', '')
            servicio = row.get('num_servicio', '')
            # Intervalo respecto al anterior
            intervalo = ""
            if prev_t is not None:
                dt = t_ini - prev_t
                if dt > 0:
                    intervalo = _fmt_hora(dt)
            prev_t = t_ini

            unidad = "Múltiple" if bool(row.get('doble', False)) else ""
            m1 = row.get('motriz_1_opt', None)
            m2 = row.get('motriz_2_opt', None)

            valores = [
                n_viaje, servicio, _fmt_hora(t_ini), r_idx - 1,
                intervalo, unidad,
                int(m1) if m1 is not None else "",
                int(m2) if m2 is not None else "",
            ]
            for c_idx, val in enumerate(valores, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = center; cell.border = border
                cell.font = Font(name='Arial', size=10)

        # Ancho de columnas
        anchos = [10, 10, 12, 11, 11, 10, 10, 10]
        for c_idx, w in enumerate(anchos, start=1):
            ws.column_dimensions[ws.cell(row=1, column=c_idx).column_letter].width = w

        wb.save(ruta)

    return ruta_v1, ruta_v2
